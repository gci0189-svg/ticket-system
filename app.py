"""
票務系統 v3.0 — 全面重新設計
左側導覽切換頁面：上傳與處理 / 標籤產生 / 簽到表 / 補印 / 歷史與設定
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
from collections import defaultdict
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════
# 頁面設定 ＆ 設計系統
# ══════════════════════════════════════════════════════════
st.set_page_config(page_title="票務系統", page_icon="🎫", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;500;700;900&family=Space+Mono:wght@400;700&display=swap');

:root {
  --primary: #E94560;
  --primary-dark: #C73652;
  --dark: #1A1A2E;
  --border: #ECECF1;
  --muted: #8A8A93;
  --bg-soft: #FAFAFC;
  --success: #2E7D32;
  --warn: #B8860B;
}

html, body, [class*="css"] { font-family: 'Noto Sans TC', sans-serif; }

/* 隱藏預設元素 */
#MainMenu, footer { visibility: hidden; }
.stDeployButton { display: none; }
header[data-testid="stHeader"] { background: transparent; }

/* ── 頂部品牌列 ── */
.brand-bar {
  background: linear-gradient(135deg, var(--dark) 0%, #0f3460 100%);
  border-radius: 16px;
  padding: 1.1rem 1.6rem;
  margin-bottom: 1.25rem;
  display: flex; align-items: center; justify-content: space-between;
}
.brand-bar h1 {
  font-family: 'Space Mono', monospace;
  color: var(--primary);
  font-size: 1.4rem; margin: 0; letter-spacing: -0.5px;
}
.brand-bar .sub { color: rgba(255,255,255,0.45); font-size: 0.8rem; margin-top: 2px; }
.brand-bar .page-pill {
  background: rgba(233,69,96,0.15); color: var(--primary);
  padding: 0.3rem 0.9rem; border-radius: 20px; font-size: 0.82rem; font-weight: 700;
}

/* ── 卡片 ── */
.card {
  background: #fff; border: 1px solid var(--border); border-radius: 14px;
  padding: 1.2rem 1.4rem; margin-bottom: 1rem;
}
.card-title {
  font-weight: 700; font-size: 0.95rem; color: var(--dark);
  margin-bottom: 0.8rem; display: flex; align-items: center; gap: 0.4rem;
}

/* ── 統計條 ── */
.stat-strip { display: flex; gap: 0.7rem; margin-bottom: 1rem; flex-wrap: wrap; }
.stat-pill {
  flex: 1; min-width: 120px; background: #fff; border: 1px solid var(--border);
  border-radius: 12px; padding: 0.7rem 1rem; text-align: center;
}
.stat-pill .num { font-family: 'Space Mono', monospace; font-size: 1.5rem; font-weight: 700; color: var(--dark); }
.stat-pill .lbl { font-size: 0.72rem; color: var(--muted); margin-top: 2px; }
.stat-pill.accent .num { color: var(--primary); }

/* ── 規則卡 ── */
.rule-box {
  background: var(--bg-soft); border: 1px solid var(--border); border-radius: 10px;
  padding: 0.7rem 0.95rem; margin-bottom: 0.5rem; font-size: 0.85rem;
}
.rule-key { color: var(--muted); font-size: 0.72rem; margin-bottom: 0.15rem; }
.rule-val { font-weight: 600; color: var(--dark); }

/* ── 群組標頭 ── */
.group-head {
  background: var(--bg-soft); border-left: 3px solid var(--primary);
  border-radius: 8px; padding: 0.5rem 0.9rem; margin: 0.7rem 0 0.5rem;
  font-weight: 700; font-size: 0.85rem; color: var(--dark);
  display: flex; justify-content: space-between; align-items: center;
}

/* ── 提示框 ── */
.note-ok   { background:#EAF6EC; border:1px solid #BFE3C4; border-radius:10px; padding:0.7rem 1rem; font-size:0.85rem; color:#1E5631; margin-bottom:0.8rem; }
.note-warn { background:#FFF6E0; border:1px solid #F3D98A; border-radius:10px; padding:0.7rem 1rem; font-size:0.85rem; color:#7A5B00; margin-bottom:0.8rem; }
.note-info { background:#EEF3FF; border:1px solid #C9D9FF; border-radius:10px; padding:0.7rem 1rem; font-size:0.85rem; color:#1F3A8A; margin-bottom:0.8rem; }

/* ── 側邊導覽 ── */
section[data-testid="stSidebar"] { background: #fff; border-right: 1px solid var(--border); }
.nav-title { font-size: 0.72rem; color: var(--muted); font-weight: 700; letter-spacing: 1px; margin: 1rem 0 0.4rem; text-transform: uppercase; }

/* dataframe / data_editor 圓角 */
[data-testid="stDataFrame"], [data-testid="stDataEditor"] { border-radius: 10px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# Google Sheets
# ══════════════════════════════════════════════════════════
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive.readonly"]

HISTORY_SHEET_ID = "1McLgRi-4haGs1orOXlCaucL6lM9E9uFgt4El8R4ZdkA"

@st.cache_resource(ttl=300)
def get_gc():
    try:
        creds = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=SCOPES)
        return gspread.authorize(creds)
    except:
        return None

def load_history(sid: str) -> set:
    gc = get_gc()
    if not gc or not sid:
        return set()
    try:
        sh = gc.open_by_key(sid)
    except gspread.exceptions.APIError as e:
        st.warning(f"歷史紀錄試算表連線失敗（{e.response.status_code}）：請確認 Sheet ID 與服務帳戶權限")
        return set()
    except Exception as e:
        st.warning(f"歷史紀錄連線失敗：{e}")
        return set()
    try:
        ws = sh.worksheet("已列印紀錄")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet("已列印紀錄", 1000, 3)
        ws.append_row(["唯一識別碼", "列印時間", "標籤內容"])
        return set()
    try:
        vals = ws.col_values(1)
        return {v for v in vals if v and v != "唯一識別碼"}
    except Exception as e:
        st.warning(f"讀取歷史紀錄失敗：{e}")
        return set()

def save_history(sid: str, keys: list, labels: list) -> bool:
    gc = get_gc()
    if not gc:
        st.error("Google Sheets 連線失敗")
        return False
    try:
        sh = gc.open_by_key(sid)
        try:
            ws = sh.worksheet("已列印紀錄")
        except:
            ws = sh.add_worksheet("已列印紀錄", 1000, 3)
            ws.append_row(["唯一識別碼", "列印時間", "標籤內容"])
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append_rows([[k, now, l] for k, l in zip(keys, labels)])
        return True
    except Exception as e:
        st.error(f"儲存失敗：{e}")
        return False

def remove_history(sid: str, keys_to_remove: list) -> bool:
    """從已列印紀錄中移除指定 key（取消歸檔用）"""
    gc = get_gc()
    if not gc:
        st.error("Google Sheets 連線失敗")
        return False
    try:
        sh = gc.open_by_key(sid)
        ws = sh.worksheet("已列印紀錄")
        all_values = ws.get_all_values()
        if not all_values:
            return True
        header = all_values[0]
        keep_rows = [header]
        removed = 0
        for row in all_values[1:]:
            if row and row[0] in keys_to_remove:
                removed += 1
                continue
            keep_rows.append(row)
        ws.clear()
        ws.append_rows(keep_rows)
        return True
    except Exception as e:
        st.error(f"取消歸檔失敗：{e}")
        return False


# ══════════════════════════════════════════════════════════
# 解析函數：LINE 會員格式（5月親子｜會員購票）
#
# 欄位對應（已與使用者確認）：
#   Row0 = 標題列
#   Row1 = 彙總計算列（跳過）
#   B欄(index=1) = 編號
#   D欄(index=3) = 姓名
#   G欄(index=6) = 日期場次
#   H欄(index=7) = 座位
#   I欄(index=8) = 張數
#   J欄(index=9) = 票價
#
# 列印條件：座位、張數、票價三欄都有值
# 合併：相同「編號」的多行 → 張數加總、取最早場次
# 唯一識別碼：{編號}_{工作表名稱}
# 標籤格式：NO.{編號} {最早場次} 貴賓｜{姓名} X {總張數}
# ══════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════
# 簽到表產生函數：LINE 會員格式
# ══════════════════════════════════════════════════════════

def _cjk_len(s):
    """計算含中文的字串顯示寬度"""
    return sum(2 if "\u4e00" <= c <= "\u9fff" or "\u3000" <= c <= "\u303f" or "\uff00" <= c <= "\uffef" else 1 for c in str(s))

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _has_value(row, idx):
    if idx >= len(row): return False
    v = row[idx]
    if v is None: return False
    if isinstance(v, (int, float)): return v != 0
    return str(v).strip() != ""

def _get(row, idx):
    if idx >= len(row): return ""
    v = row[idx]
    if v is None: return ""
    if isinstance(v, datetime): return v.strftime("%Y/%m/%d")
    return str(v).strip()


# ══════════════════════════════════════════════════════════
# 把網頁上的編輯寫回 xlsx
# ══════════════════════════════════════════════════════════
def apply_edits_to_xlsx(file_bytes: bytes, sheet_name: str, edits: dict, all_tickets: list) -> bytes:
    """
    把 edits 寫回原始 xlsx 的對應欄位。
    只改 LINE 會員格式（姓名 D欄、張數 I欄、電話 E欄、座位 H欄）。
    回傳修改後的 xlsx bytes。
    """
    from openpyxl import load_workbook
    from io import BytesIO
    import re

    wb = load_workbook(BytesIO(file_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到工作表：{sheet_name}")
    ws = wb[sheet_name]

    # 建立 編號 → row_number 對應（從 Row3 開始）
    id_to_rows = {}
    for row in ws.iter_rows(min_row=3):
        cell_id = str(row[1].value or "").strip()
        id_digits = re.sub(r"[^0-9]", "", cell_id)
        if id_digits:
            id_to_rows.setdefault(id_digits, []).append(row[0].row)

    # 建立 key → ticket 對應
    key_to_ticket = {t["key"]: t for t in all_tickets}

    COL_NAME  = 4   # D（1-based）
    COL_TEL   = 5   # E
    COL_SEAT  = 8   # H
    COL_COUNT = 9   # I

    for key, ed in edits.items():
        ticket = key_to_ticket.get(key)
        if not ticket:
            continue
        row_id = ticket["id"]
        row_nums = id_to_rows.get(row_id, [])
        if not row_nums:
            continue
        # 只改第一行（主行）
        first_row = min(row_nums)
        if "name" in ed:
            ws.cell(row=first_row, column=COL_NAME).value = ed["name"]
        if "tel" in ed:
            ws.cell(row=first_row, column=COL_TEL).value = ed["tel"]
        if "seats" in ed:
            ws.cell(row=first_row, column=COL_SEAT).value = ed["seats"]
        if "count" in ed:
            # 如果只有一行就直接改，多行的話只改第一行
            ws.cell(row=first_row, column=COL_COUNT).value = ed["count"]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generate_signin_excel(file_bytes: bytes, sheet_name: str, show_name: str) -> bytes:
    """
    從報名表 xlsx 產生簽到表 Excel。
    目前支援 LINE 會員格式（5月親子｜會員購票）。
    回傳 Excel bytes。
    """
    wb_src = load_workbook(BytesIO(file_bytes), read_only=True)
    ws_src = wb_src[sheet_name]
    rows = list(ws_src.iter_rows(values_only=True))

    COL_ID=1; COL_SNS=2; COL_NAME=3; COL_TEL=4
    COL_DATE=6; COL_SEAT=7; COL_COUNT=8; COL_PRICE=9

    merged = {}
    last_id = None
    for i, row in enumerate(rows):
        if i < 2: continue
        if not _has_value(row, COL_SEAT): continue
        if not _has_value(row, COL_COUNT): continue
        if not _has_value(row, COL_PRICE): continue

        count_raw = _get(row, COL_COUNT)
        count_clean = re.sub(r"[^0-9]", "", count_raw)
        if not count_clean: continue
        count = int(count_clean)
        if count <= 0: continue

        row_id = _get(row, COL_ID)
        id_digits = re.sub(r"[^0-9]", "", row_id)
        if id_digits: last_id = id_digits
        if not last_id: continue

        name = _get(row, COL_NAME)
        sns  = _get(row, COL_SNS)
        tel  = _get(row, COL_TEL)
        date_raw = _get(row, COL_DATE)

        if last_id not in merged:
            m = re.search(r"\d{4}[/\-](\d{1,2})[/\-](\d{1,2})", date_raw)
            month = int(m.group(1)) if m else 99
            day   = int(m.group(2)) if m else 99
            sess_ord = 1 if any(x in date_raw for x in ["14:30","下午"]) else 0
            sess_str = "下午" if sess_ord else "上午"
            wday_m = re.search(r"[(（](.)[ )）]", date_raw)
            wday = wday_m.group(1) if wday_m else ""
            time_str = "14:30" if sess_ord else "10:30"
            full_date = f"2026/{month:02d}/{day:02d}（{wday}）{time_str}"
            merged[last_id] = {
                "id": last_id, "name": name, "sns": sns, "tel": tel,
                "seats": [], "total": 0,
                "sort": (month, day, sess_ord),
                "session": f"{month}/{day} {sess_str}",
                "full_date": full_date
            }
        else:
            if not merged[last_id]["name"] and name: merged[last_id]["name"] = name
            if not merged[last_id]["sns"] and sns:   merged[last_id]["sns"] = sns
            if not merged[last_id]["tel"] and tel:   merged[last_id]["tel"] = tel

        seat = _get(row, COL_SEAT)
        if seat:
            for s in re.split(r"[\n\r]+", seat):
                s = s.strip()
                if s: merged[last_id]["seats"].append(s)
        merged[last_id]["total"] += count

    sorted_data = sorted(merged.values(), key=lambda x: (x["sort"], int(x["id"])))
    sessions = defaultdict(list)
    for r in sorted_data:
        sessions[r["session"]].append(r)

    # 樣式
    HDR_FONT = Font(name="Arial", bold=True, size=11)
    SUB_FILL = PatternFill("solid", start_color="E94560")
    SUB_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    ALT_FILL  = PatternFill("solid", start_color="FFF5F7")
    TOT_FILL  = PatternFill("solid", start_color="EEEEEE")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for session_display, sess_rows in sessions.items():
        safe_name = session_display.replace("/", "-")
        ws = wb_out.create_sheet(title=safe_name)
        first = sess_rows[0]

        # Row1：A1=日期，C1=劇名（比照舊版）
        ws.row_dimensions[1].height = 22
        ws["A1"] = first["full_date"]
        ws["A1"].font = HDR_FONT
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["C1"] = show_name
        ws["C1"].font = HDR_FONT
        ws["C1"].alignment = Alignment(horizontal="left", vertical="center")

        # Row2：欄位標題
        ws.row_dimensions[2].height = 20
        col_headers = ["編號", "社群帳號", "姓名", "電話", "座位", "張數", "領取簽名"]
        for col, h in enumerate(col_headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = SUB_FONT; c.fill = SUB_FILL
            c.alignment = CENTER; c.border = _thin_border()

        # 資料列
        for idx, r in enumerate(sess_rows):
            row_num = idx + 3
            fill = ALT_FILL if idx % 2 == 1 else None
            seat_str = "　".join(r["seats"])
            values = [r["id"], r["sns"], r["name"], r["tel"], seat_str, r["total"], ""]
            aligns = [CENTER, LEFT, CENTER, CENTER, LEFT, CENTER, CENTER]
            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.font = BODY_FONT; c.alignment = aln; c.border = _thin_border()
                if fill: c.fill = fill

        # 合計列
        total_row = len(sess_rows) + 3
        ws.merge_cells(f"A{total_row}:E{total_row}")
        c = ws.cell(row=total_row, column=1, value=f"合計：{len(sess_rows)} 人")
        c.font = BOLD_FONT; c.alignment = CENTER; c.fill = TOT_FILL; c.border = _thin_border()
        tc = ws.cell(row=total_row, column=6, value=f"=SUM(F3:F{total_row-1})")
        tc.font = BOLD_FONT; tc.alignment = CENTER; tc.fill = TOT_FILL; tc.border = _thin_border()
        ws.cell(row=total_row, column=7).fill = TOT_FILL
        ws.cell(row=total_row, column=7).border = _thin_border()

        # 自動欄寬
        col_data = {
            1: [r["id"] for r in sess_rows] + ["編號"],
            2: [r["sns"] for r in sess_rows] + ["社群帳號"],
            3: [r["name"] for r in sess_rows] + ["姓名"],
            4: [r["tel"] for r in sess_rows] + ["電話"],
            5: ["　".join(r["seats"]) for r in sess_rows] + ["座位"],
            6: [r["total"] for r in sess_rows] + ["張數"],
            7: ["領取簽名"],
        }
        min_ws_map = {1:6, 2:12, 3:8, 4:14, 5:20, 6:6, 7:12}
        max_ws_map = {1:8, 2:25, 3:12, 4:16, 5:45, 6:8, 7:15}
        for col_idx, values in col_data.items():
            max_w = max((_cjk_len(v) for v in values if v), default=min_ws_map[col_idx])
            w = min(max(max_w + 2, min_ws_map[col_idx]), max_ws_map[col_idx])
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # 自動列高
        col5_w = ws.column_dimensions["E"].width
        for idx, r in enumerate(sess_rows):
            row_num = idx + 3
            seat_str = "　".join(r["seats"])
            lines = max(1, int(_cjk_len(seat_str) / max(col5_w, 1)) + 1)
            ws.row_dimensions[row_num].height = max(18, lines * 16)

        ws.freeze_panes = "A3"

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue()

def parse_session(raw: str):
    """解析日期場次字串，回傳 (排序key, 顯示文字)
    支援格式：
      2026/05/24(日)14:30、5/24 下午、2026-05-24 10:30
    """
    if not raw:
        return (99, 99, 99), ""
    raw = str(raw)
    # 優先嘗試 年/月/日 格式（避免誤把年份當月份）
    m = re.search(r'\d{4}[/\-](\d{1,2})[/\-](\d{1,2})', raw)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
    else:
        m = re.search(r'(\d{1,2})[/／\-月](\d{1,2})', raw)
        month = int(m.group(1)) if m else 99
        day   = int(m.group(2)) if m else 99
    if any(x in raw for x in ["14:30","15:00","下午","PM","pm"]):
        sess_str, sess_ord = "下午", 1
    elif any(x in raw for x in ["10:30","10:00","上午","AM","am"]):
        sess_str, sess_ord = "上午", 0
    else:
        sess_str, sess_ord = "", 0
    display = f"{month}/{day} {sess_str}".strip()
    return (month, day, sess_ord), display



# ══════════════════════════════════════════════════════════
# 解析函數：貴賓印製標籤版 ＆ 社福印製標籤版
#
# 欄位對應（已確認）：
#   Row0 = 劇名（略過）
#   Row1 = 標題列：編號、日期、姓名、座位（貴賓）/張數（社福）、張數
#   Row2 起 = 資料
#   A欄(0) = 編號
#   B欄(1) = 日期場次
#   C欄(2) = 姓名（已含「貴賓｜單位｜姓名」格式，直接使用）
#   D欄(3) = 座位（貴賓）或 張數（社福）
#   E欄(4) = 張數（貴賓）/ 無（社福）
#
# 標籤格式：{場次} {姓名欄原樣} X {張數}
# 唯一識別碼：{編號}_{工作表名稱}
# ══════════════════════════════════════════════════════════

def parse_label_sheet(df: pd.DataFrame, sheet_name: str, history_set: set):
    """
    解析貴賓印製標籤版 / 社福印製標籤版。
    Row0=劇名, Row1=標題, Row2起=資料
    """
    rows = df.values.tolist()
    results  = []
    warnings = []

    # 判斷是貴賓（有座位欄）還是社福（無座位欄）
    row1_str = " ".join(str(c) for c in rows[1]) if len(rows) > 1 else ""
    has_seat_col = "座位" in row1_str
    # 貴賓：A=編號 B=日期 C=姓名 D=座位 E=張數
    # 社福：A=編號 B=日期 C=姓名 D=張數
    COL_ID    = 0
    COL_DATE  = 1
    COL_NAME  = 2
    COL_COUNT = 4 if has_seat_col else 3

    for i, row in enumerate(rows):
        if i < 2: continue  # 跳過劇名列和標題列

        def get(idx):
            if idx >= len(row): return ""
            v = row[idx]
            if v is None: return ""
            if isinstance(v, __import__('datetime').datetime): return v.strftime("%Y/%m/%d")
            return str(v).strip()

        row_id   = re.sub(r"[^0-9]", "", get(COL_ID))
        name     = get(COL_NAME)
        date_raw = get(COL_DATE)
        count_raw = get(COL_COUNT)

        if not row_id or not name or not count_raw:
            continue

        count_clean = re.sub(r"[^0-9]", "", count_raw)
        if not count_clean: continue
        count = int(count_clean)
        if count <= 0: continue

        sort_key, display = parse_session(date_raw)
        key = f"{row_id}_{sheet_name}"
        label = f"{display} {name} X {count}"

        entry = {
            "key":              key,
            "id":               row_id,
            "name":             name,
            "sheet":            sheet_name,
            "total_count":      count,
            "count":            count,
            "earliest_sort":    sort_key,
            "earliest_display": display,
            "label":            label,
            "sns":              "",
            "tel":              "",
            "seats":            [],
            "is_new":           key not in history_set,
            "fmt":              "label",
        }

        if entry["is_new"]:
            results.append(entry)
        else:
            entry_copy = dict(entry)
            results.append(entry_copy)

    # 分開 new / skipped
    tickets = [r for r in results if r["is_new"]]
    skipped = [r for r in results if not r["is_new"]]
    tickets.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
    skipped.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
    return tickets, skipped, warnings

def parse_member_sheet(df: pd.DataFrame, sheet_name: str, history_set: set):
    """
    解析 LINE 會員格式工作表。
    回傳 (tickets, skipped, warnings)
    """
    rows = df.values.tolist()

    COL_ID    = 1   # B：編號
    COL_NAME  = 3   # D：姓名
    COL_DATE  = 6   # G：日期場次
    COL_SEAT  = 7   # H：座位
    COL_COUNT = 8   # I：張數
    COL_PRICE = 9   # J：票價

    def get(row, col):
        if col >= len(row):
            return ""
        v = row[col]
        return str(v).strip() if v is not None else ""

    merged   = {}   # unique_key → dict
    warnings = []
    last_id  = None  # 追蹤最近一個有效編號

    for i, row in enumerate(rows):
        if i < 2:
            continue  # 跳過 Row0（標題）和 Row1（彙總）

        seat      = get(row, COL_SEAT)
        count_raw = get(row, COL_COUNT)
        price     = get(row, COL_PRICE)

        # 跳過 A欄有「已印票」標記的列
        a_val = str(row[0] or "").strip() if len(row) > 0 else ""
        if "已印票" in a_val:
            continue

        # 列印條件：三欄都要有值
        if not seat or not count_raw or not price:
            continue

        # 張數解析
        count_clean = re.sub(r'[^0-9]', '', count_raw)
        if not count_clean:
            warnings.append(f"第 {i+1} 行：張數無法解析（{count_raw!r}）")
            continue
        count = int(count_clean)
        if count <= 0:
            continue

        # 票價必須像數字
        if not re.search(r'\d', price):
            continue

        row_id   = get(row, COL_ID)
        name     = get(row, COL_NAME)
        date_raw = get(row, COL_DATE)

        # 更新 last_id
        id_digits = re.sub(r'[^0-9]', '', row_id)
        if id_digits:
            last_id = id_digits
        
        if not last_id:
            warnings.append(f"第 {i+1} 行：找不到編號，已略過")
            continue

        key = f"{last_id}_{sheet_name}"
        sort_key, display = parse_session(date_raw)

        if key not in merged:
            # C欄（社群帳號）若含「樓」「排」表示誤填了座位資訊，視為空白
            raw_sns = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            sns_val = "" if any(x in raw_sns for x in ["樓", "排"]) else raw_sns
            merged[key] = {
                "key":              key,
                "id":               last_id,
                "name":             name or "(未填姓名)",
                "sns":              sns_val,
                "tel":              str(row[4]).strip() if len(row) > 4 and row[4] else "",
                "seats":            [],
                "sheet":            sheet_name,
                "total_count":      0,
                "earliest_sort":    sort_key,
                "earliest_display": display,
                "is_new":           key not in history_set,
            }
        else:
            # 更新最早場次
            if sort_key < merged[key]["earliest_sort"]:
                merged[key]["earliest_sort"]   = sort_key
                merged[key]["earliest_display"] = display
            # 補姓名（跨行的附屬行可能是空的）
            if (not merged[key]["name"] or merged[key]["name"] == "(未填姓名)") and name:
                merged[key]["name"] = name

        # 收集座位（同格內換行也分開）
        raw_seat = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        if raw_seat:
            import re as _re
            for s in _re.split(r"[\n\r]+", raw_seat):
                s = s.strip()
                if s: merged[key]["seats"].append(s)
        merged[key]["total_count"] += count

    # 產生標籤
    tickets = []
    skipped = []
    for key, info in merged.items():
        label = (f"NO.{info['id']} {info['earliest_display']} "
                 f"貴賓｜{info['name']} X {info['total_count']}")
        entry = {**info, "label": label, "count": info["total_count"],
                  "sns": info.get("sns",""), "tel": info.get("tel",""),
                  "seats": info.get("seats",[]), "fmt": "member"}
        if info["is_new"]:
            tickets.append(entry)
        else:
            skipped.append(entry)

    tickets.sort(key=lambda x: x["earliest_sort"])
    skipped.sort(key=lambda x: x["earliest_sort"])
    return tickets, skipped, warnings



# ══════════════════════════════════════════════════════════
# 解析函數：台北台語格式（場次橫向展開，方案A合併）
#
# Row0 有「詢問時間」可識別
# col1=編號, col5=社群帳號, col6=姓名, col7=電話, col9=主日期
# col10=座位, col11~14=各場次張數, col15=張數(SUM), col22=現場取票
# 列印條件：col22 非空（現場取票）
# 合併：相同編號 → 各場次張數加總、取最早場次（方案A）
# ══════════════════════════════════════════════════════════
def parse_taibei_sheet(df: pd.DataFrame, sheet_name: str, history_set: set):
    rows = df.values.tolist()
    warnings_list = []

    COL_ID     = 1
    COL_SNS    = 5
    COL_NAME   = 6
    COL_TEL    = 7
    COL_DATE   = 9   # 主日期（下拉選的）
    COL_SEAT   = 10
    COL_PICKUP = 22  # 現場取票

    # 動態偵測場次欄：Row0 中 col11~col20 含「/」的欄位
    if len(rows) < 1:
        return [], [], []
    header = [str(c or "").strip() for c in rows[0]]
    session_cols = []
    for ci in range(11, min(21, len(header))):
        h = header[ci].replace("\n", "")
        if "/" in h and any(t in h for t in ["10:30","14:30","10","14"]):
            session_cols.append((ci, h.replace("\n", " ")))

    def get_val(row, idx):
        if idx >= len(row): return ""
        v = row[idx]
        if v is None: return ""
        if isinstance(v, __import__("datetime").datetime): return v.strftime("%Y/%m/%d")
        return str(v).strip()

    merged = {}
    last_id = None

    for i, row in enumerate(rows):
        if i < 2: continue

        # 現場取票欄
        pickup = get_val(row, COL_PICKUP).upper()
        if not pickup:
            continue  # 空白代表不需現場取票

        row_id_raw = get_val(row, COL_ID)
        id_digits  = re.sub(r"[^0-9]", "", row_id_raw)
        if id_digits: last_id = id_digits
        if not last_id: continue

        name = get_val(row, COL_NAME)
        sns  = get_val(row, COL_SNS)
        tel  = get_val(row, COL_TEL)
        seat = get_val(row, COL_SEAT)
        date_raw = get_val(row, COL_DATE)

        # 遍歷場次欄，累計有值的張數
        total = 0
        earliest_sort   = (99, 99, 99)
        earliest_display = ""
        for ci, sess_label in session_cols:
            val = get_val(row, ci)
            count_clean = re.sub(r"[^0-9]", "", val)
            if not count_clean: continue
            count = int(count_clean)
            if count <= 0: continue
            # 從場次標題解析日期
            sort_key, disp = parse_session(sess_label)
            if sort_key < earliest_sort:
                earliest_sort    = sort_key
                earliest_display = disp
            total += count

        # 若場次欄都沒值，嘗試從主日期 + 張數欄讀
        if total == 0:
            # col15 是 SUM 公式，pandas 讀為字串；嘗試 col11~14 SUM
            for ci in range(11, 15):
                v = get_val(row, ci)
                c = re.sub(r"[^0-9]", "", v)
                if c: total += int(c)
            if total == 0: continue
            if not earliest_display:
                earliest_sort, earliest_display = parse_session(date_raw)

        if not earliest_display:
            earliest_sort, earliest_display = parse_session(date_raw)

        key = f"{last_id}_{sheet_name}"
        if key not in merged:
            # C欄社群帳號誤填座位防呆
            raw_sns = sns
            if any(x in raw_sns for x in ["樓","排"]): raw_sns = ""
            merged[key] = {
                "key": key, "id": last_id, "name": name or "(未填姓名)",
                "sns": raw_sns, "tel": tel, "seats": [seat] if seat else [],
                "sheet": sheet_name, "total_count": 0,
                "earliest_sort": earliest_sort, "earliest_display": earliest_display,
                "is_new": key not in history_set,
            }
        else:
            if earliest_sort < merged[key]["earliest_sort"]:
                merged[key]["earliest_sort"]    = earliest_sort
                merged[key]["earliest_display"] = earliest_display
            if not merged[key]["name"] or merged[key]["name"] == "(未填姓名)":
                if name: merged[key]["name"] = name
            if seat and seat not in merged[key]["seats"]:
                merged[key]["seats"].append(seat)

        merged[key]["total_count"] += total

    tickets, skipped = [], []
    for key, info in merged.items():
        label = (f"NO.{info['id']} {info['earliest_display']} "
                 f"貴賓｜{info['name']} X {info['total_count']}")
        entry = {**info, "label": label, "count": info["total_count"],
                 "sns": info.get("sns",""), "tel": info.get("tel",""),
                 "seats": info.get("seats",[]), "fmt": "member"}
        if info["is_new"]: tickets.append(entry)
        else: skipped.append(entry)

    tickets.sort(key=lambda x: x["earliest_sort"])
    skipped.sort(key=lambda x: x["earliest_sort"])
    return tickets, skipped, warnings_list


# ══════════════════════════════════════════════════════════
# 解析函數：其他需求格式（場次橫向，無流水編號）
#
# col5=社群帳號|公司, col6=姓名, col7=電話, col9=場地
# col10=主日期, col11=座位
# col12=宜蘭場貴賓, col13=台北場貴賓, col14=台北場購票
# col21=現場取票
# 列印條件：col21 非空
# 唯一識別碼：{姓名}_{場次}_{工作表名稱}
# 標籤：{場次} 貴賓｜{公司}｜{姓名} X {張數}
# ══════════════════════════════════════════════════════════
def parse_other_sheet(df: pd.DataFrame, sheet_name: str, history_set: set):
    rows = df.values.tolist()
    warnings_list = []

    COL_COMPANY = 5
    COL_NAME    = 6
    COL_TEL     = 7
    COL_DATE    = 10  # 主日期
    COL_SEAT    = 11
    COL_PICKUP  = 21

    # 場次欄：col12, col13, col14
    SESSION_COLS = [12, 13, 14]

    def get_val(row, idx):
        if idx >= len(row): return ""
        v = row[idx]
        if v is None: return ""
        if isinstance(v, __import__("datetime").datetime): return v.strftime("%Y/%m/%d")
        return str(v).strip()

    # Row0 的場次欄標題作為場次顯示名稱
    header = [str(c or "").strip() for c in rows[0]] if rows else []
    sess_labels = {ci: header[ci] if ci < len(header) else f"場次{ci}" for ci in SESSION_COLS}

    tickets, skipped = [], []

    for i, row in enumerate(rows):
        if i < 2: continue

        # 跳過群組標題列（A欄有「貴賓」「社福」等）
        a_val = get_val(row, 0)
        if a_val and not re.search(r"\d", get_val(row, COL_NAME) + get_val(row, COL_COMPANY)):
            if not get_val(row, COL_NAME): continue

        # 現場取票
        pickup = get_val(row, COL_PICKUP)
        if not pickup: continue

        name    = get_val(row, COL_NAME)
        company = get_val(row, COL_COMPANY)
        tel     = get_val(row, COL_TEL)
        seat    = get_val(row, COL_SEAT)
        date_raw = get_val(row, COL_DATE)

        if not name: continue

        # 每個有值的場次欄各產生一筆（但同場次合併）
        for ci in SESSION_COLS:
            val = get_val(row, ci)
            count_clean = re.sub(r"[^0-9]", "", val)
            if not count_clean: continue
            count = int(count_clean)
            if count <= 0: continue

            # 場次：優先用欄標題，fallback 用主日期
            sess_raw = sess_labels.get(ci, "")
            sort_key, display = parse_session(sess_raw)
            if display == "99/99":
                sort_key, display = parse_session(date_raw)

            # key：姓名+場次+工作表
            safe = re.sub(r"\s", "", f"{name}_{display}_{sheet_name}")
            key  = safe[:80]

            label_name = f"{company}｜{name}" if company else name
            label = f"{display} 貴賓｜{label_name} X {count}"

            entry = {
                "key": key, "id": str(i), "name": name,
                "sns": "", "tel": tel, "seats": [seat] if seat else [],
                "sheet": sheet_name, "total_count": count, "count": count,
                "earliest_sort": sort_key, "earliest_display": display,
                "label": label, "fmt": "label",
                "is_new": key not in history_set,
            }
            if entry["is_new"]: tickets.append(entry)
            else: skipped.append(entry)

    tickets.sort(key=lambda x: x["earliest_sort"])
    skipped.sort(key=lambda x: x["earliest_sort"])
    return tickets, skipped, warnings_list

def group_by_session(tickets: list) -> dict:
    grouped = {}
    for t in tickets:
        d = t["earliest_display"]
        grouped.setdefault(d, []).append(t)
    return dict(sorted(grouped.items(), key=lambda x: x[1][0]["earliest_sort"]))


# ══════════════════════════════════════════════════════════
# Session State
# ══════════════════════════════════════════════════════════
defaults = {
    "page":                "upload",
    "history_sid":         HISTORY_SHEET_ID,
    "history_set":         set(),
    "raw_sheets":          {},
    "selected_sheets":     [],
    "rule_confirmed":      False,
    "tickets":             [],
    "skipped":             [],
    "warnings":            [],
    "checked_keys":        set(),
    "edits":               {},
    "uploaded_file_bytes": None,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

if not st.session_state.history_set:
    st.session_state.history_set = load_history(HISTORY_SHEET_ID)


# ══════════════════════════════════════════════════════════
# 共用：標籤文字重組（套用編輯後的姓名/張數）
# ══════════════════════════════════════════════════════════
def make_current_label(t: dict) -> str:
    ed = st.session_state.edits.get(t["key"], {})
    name  = ed.get("name",  t["name"])
    count = ed.get("count", t["count"])
    if t.get("fmt") == "member":
        parts = t["label"].split("貴賓｜")
        if len(parts) == 2:
            return f"{parts[0]}貴賓｜{name} X {count}"
        return f"NO.{t['id']} {t['earliest_display']} 貴賓｜{name} X {count}"
    return f"{t['earliest_display']} {name} X {count}"


# ══════════════════════════════════════════════════════════
# 側邊導覽
# ══════════════════════════════════════════════════════════
PAGES = {
    "upload": "📤 上傳與處理",
    "labels": "🎫 標籤產生",
    "signin": "📋 簽到表",
    "reprint": "🖨️ 補印",
    "settings": "⚙️ 歷史與設定",
}

with st.sidebar:
    st.markdown('<div class="nav-title">功能選單</div>', unsafe_allow_html=True)
    for key, label in PAGES.items():
        is_current = st.session_state.page == key
        if st.button(label, key=f"nav_{key}", use_container_width=True,
                     type="primary" if is_current else "secondary"):
            st.session_state.page = key
            st.rerun()

    st.markdown('<div class="nav-title">即時狀態</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="stat-pill" style="margin-bottom:0.5rem;">'
        f'<div class="num">{len(st.session_state.history_set)}</div>'
        f'<div class="lbl">已列印歷史筆數</div></div>',
        unsafe_allow_html=True
    )
    if st.session_state.tickets:
        n_checked = len(st.session_state.checked_keys)
        st.markdown(
            f'<div class="stat-pill accent">'
            f'<div class="num">{n_checked}/{len(st.session_state.tickets)}</div>'
            f'<div class="lbl">本次已勾選</div></div>',
            unsafe_allow_html=True
        )

# 頂部品牌列
st.markdown(f"""
<div class="brand-bar">
  <div>
    <h1>🎫 票務系統</h1>
    <div class="sub">上傳報名表 → 確認規則 → 產生標籤 → 簽到表 → 補印</div>
  </div>
  <div class="page-pill">{PAGES[st.session_state.page]}</div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# 頁面：📤 上傳與處理
# ══════════════════════════════════════════════════════════
if st.session_state.page == "upload":

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">① 上傳 xlsx 報名表</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("選取或拖曳 xlsx 檔案", type=["xlsx"], label_visibility="collapsed")
    if uploaded:
        try:
            file_bytes = uploaded.read()
            st.session_state.uploaded_file_bytes = file_bytes
            xls = pd.ExcelFile(BytesIO(file_bytes))
            raw = {}
            for sname in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sname, header=None, dtype=str)
                raw[sname] = df.fillna("")
            if set(raw.keys()) != set(st.session_state.raw_sheets.keys()):
                st.session_state.raw_sheets      = raw
                st.session_state.selected_sheets = []
                st.session_state.rule_confirmed  = False
                st.session_state.tickets         = []
                st.session_state.skipped         = []
                st.session_state.warnings        = []
                st.session_state.checked_keys    = set()
                st.session_state.edits           = {}
            st.markdown(f'<div class="note-ok">✅ 已載入：{uploaded.name}，共 {len(raw)} 個工作表</div>', unsafe_allow_html=True)
        except Exception as e:
            st.error(f"讀取失敗：{e}")
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.raw_sheets:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">② 預覽 ＆ 選擇工作表（可多選，將合併處理）</div>', unsafe_allow_html=True)

        sheet_names = list(st.session_state.raw_sheets.keys())
        preview_sheet = st.selectbox("預覽工作表內容", options=sheet_names, key="preview_select")
        st.dataframe(st.session_state.raw_sheets[preview_sheet].iloc[:8, :12],
                     use_container_width=True, height=220)

        st.markdown("<br>", unsafe_allow_html=True)
        selected_sheets = st.multiselect(
            "選擇要產生標籤的工作表",
            options=sheet_names,
            default=st.session_state.selected_sheets or []
        )
        col_sel, col_btn = st.columns([4, 1])
        with col_sel:
            if selected_sheets:
                st.caption(f"已選 {len(selected_sheets)} 張：{'、'.join(selected_sheets)}")
        with col_btn:
            if st.button("✅ 確認選擇", type="primary", use_container_width=True, disabled=not selected_sheets):
                st.session_state.selected_sheets = selected_sheets
                st.session_state.rule_confirmed  = False
                st.session_state.tickets         = []
                st.session_state.skipped         = []
                st.session_state.warnings        = []
                st.session_state.checked_keys    = set()
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.selected_sheets:
        sname = st.session_state.selected_sheets[0]
        df    = st.session_state.raw_sheets[sname]

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="card-title">③ 確認解析規則：{"、".join(st.session_state.selected_sheets)}</div>', unsafe_allow_html=True)

        row0_str = " ".join(str(c) for c in df.iloc[0].tolist()) if len(df) > 0 else ""
        row1_str = " ".join(str(c) for c in df.iloc[1].tolist()) if len(df) > 1 else ""

        is_member        = ("姓名" in row0_str and "張數" in row0_str and "座位" in row0_str and "詢問時間" not in row0_str)
        is_taibei        = ("姓名" in row0_str and "座位" in row0_str and "詢問時間" in row0_str and "場地" not in row0_str)
        is_other         = ("社群帳號｜公司" in row0_str and "場地" in row0_str)
        is_label_vip     = ("編號" in row1_str and "日期" in row1_str and "張數" in row1_str and "座位" in row1_str)
        is_label_welfare = ("編號" in row1_str and "日期" in row1_str and "張數" in row1_str and "座位" not in row1_str and "姓名" not in row0_str)

        if is_member:
            st.markdown('<div class="note-ok">✅ 偵測到：<strong>LINE 會員格式</strong></div>', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            with c1:
                for k, v in [("標題列位置","第 1 行"), ("跳過","第 2 行（彙總列）"),
                             ("列印條件","座位＋張數＋票價三欄都有值"), ("合併邏輯","相同編號多行→張數加總，取最早場次")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)
            with c2:
                for k, v in [("唯一識別碼","編號＋工作表名稱"), ("標籤格式","NO.{編號} {場次} 貴賓｜{姓名} X {張數}"),
                             ("排序","依最早場次"), ("輸出","每場次／工作表分組")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✅ 確認，前往標籤產生", type="primary", use_container_width=True):
                with st.spinner("解析中..."):
                    all_t, all_s, all_w = [], [], []
                    for sh in st.session_state.selected_sheets:
                        t, s, w = parse_member_sheet(st.session_state.raw_sheets[sh], sh, st.session_state.history_set)
                        all_t.extend(t); all_s.extend(s); all_w.extend(w)
                    all_t.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
                st.session_state.tickets        = all_t
                st.session_state.skipped        = all_s
                st.session_state.warnings       = all_w
                st.session_state.rule_confirmed = True
                st.session_state.checked_keys   = set()
                st.session_state.page           = "labels"
                st.rerun()

        elif is_taibei:
            st.markdown('<div class="note-ok">✅ 偵測到：<strong>台北台語格式</strong>（場次橫向展開）</div>', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            with c1:
                for k, v in [("標題列位置","第 1 行"), ("跳過","第 2 行（彙總列）"),
                             ("列印條件","現場取票欄非空"), ("合併邏輯","相同編號多場次→張數加總，取最早場次")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)
            with c2:
                for k, v in [("唯一識別碼","編號＋工作表名稱"), ("標籤格式","NO.{編號} {最早場次} 貴賓｜{姓名} X {總張數}"),
                             ("排序","依最早場次"), ("輸出","每場次分組")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✅ 確認，前往標籤產生", type="primary", use_container_width=True):
                with st.spinner("解析中..."):
                    all_t, all_s, all_w = [], [], []
                    for sh in st.session_state.selected_sheets:
                        t, s, w = parse_taibei_sheet(st.session_state.raw_sheets[sh], sh, st.session_state.history_set)
                        all_t.extend(t); all_s.extend(s); all_w.extend(w)
                    all_t.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
                st.session_state.tickets = all_t; st.session_state.skipped = all_s
                st.session_state.warnings = all_w; st.session_state.rule_confirmed = True
                st.session_state.checked_keys = set(); st.session_state.page = "labels"
                st.rerun()

        elif is_other:
            st.markdown('<div class="note-ok">✅ 偵測到：<strong>其他需求格式</strong>（宜蘭/台北場貴賓）</div>', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            with c1:
                for k, v in [("標題列位置","第 1 行"), ("資料從","第 3 行開始"),
                             ("列印條件","現場取票欄非空"), ("合併邏輯","不合併，每場次各一筆")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)
            with c2:
                for k, v in [("唯一識別碼","姓名＋場次＋工作表名稱"), ("標籤格式","{場次} 貴賓｜{公司}｜{姓名} X {張數}"),
                             ("場次欄","col12=宜蘭場貴賓、col13=台北場貴賓、col14=台北場購票"), ("輸出","每場次分組")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✅ 確認，前往標籤產生", type="primary", use_container_width=True):
                with st.spinner("解析中..."):
                    all_t, all_s, all_w = [], [], []
                    for sh in st.session_state.selected_sheets:
                        t, s, w = parse_other_sheet(st.session_state.raw_sheets[sh], sh, st.session_state.history_set)
                        all_t.extend(t); all_s.extend(s); all_w.extend(w)
                    all_t.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
                st.session_state.tickets = all_t; st.session_state.skipped = all_s
                st.session_state.warnings = all_w; st.session_state.rule_confirmed = True
                st.session_state.checked_keys = set(); st.session_state.page = "labels"
                st.rerun()

        elif is_label_vip or is_label_welfare:
            fmt_name = "貴賓印製標籤版" if is_label_vip else "社福印製標籤版"
            st.markdown(f'<div class="note-ok">✅ 偵測到：<strong>{fmt_name}</strong></div>', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            with c1:
                for k, v in [("標題列位置","第 1 行"), ("資料從","第 3 行開始"),
                             ("列印條件","張數有值即列印"), ("合併邏輯","不合併，每行各自一筆")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)
            with c2:
                for k, v in [("唯一識別碼","編號＋工作表名稱"), ("標籤格式","{場次} {姓名欄原樣} X {張數}"),
                             ("排序","依日期場次"), ("輸出","每場次／工作表分組")]:
                    st.markdown(f'<div class="rule-box"><div class="rule-key">{k}</div><div class="rule-val">{v}</div></div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✅ 確認，前往標籤產生", type="primary", use_container_width=True):
                with st.spinner("解析中..."):
                    all_t, all_s, all_w = [], [], []
                    for sh in st.session_state.selected_sheets:
                        t, s, w = parse_label_sheet(st.session_state.raw_sheets[sh], sh, st.session_state.history_set)
                        all_t.extend(t); all_s.extend(s); all_w.extend(w)
                    all_t.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
                st.session_state.tickets        = all_t
                st.session_state.skipped        = all_s
                st.session_state.warnings       = all_w
                st.session_state.rule_confirmed = True
                st.session_state.checked_keys   = set()
                st.session_state.page           = "labels"
                st.rerun()
        else:
            st.markdown("""
            <div class="note-warn">
            ⚠️ <strong>這張工作表的格式目前尚未設定解析規則。</strong><br><br>
            請把這份 xlsx 傳給管理員確認格式後更新程式，才能安全產生標籤。<br>
            <strong>現場取票時間寶貴，絕對不能猜測格式。</strong>
            </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# 頁面：🎫 標籤產生
# ══════════════════════════════════════════════════════════
elif st.session_state.page == "labels":

    if not st.session_state.rule_confirmed or not st.session_state.tickets and not st.session_state.skipped:
        st.markdown('<div class="note-info">請先到「📤 上傳與處理」完成上傳與規則確認。</div>', unsafe_allow_html=True)
    else:
        tickets  = st.session_state.tickets
        skipped  = st.session_state.skipped
        warnings = st.session_state.warnings

        if warnings:
            with st.expander(f"⚠️ {len(warnings)} 筆資料需要注意"):
                for w in warnings:
                    st.markdown(f"- {w}")

        total_count = sum(t["count"] for t in tickets)
        n_sessions  = len(set(t["earliest_display"] for t in tickets))
        n_checked   = len(st.session_state.checked_keys)

        st.markdown(f"""
        <div class="stat-strip">
          <div class="stat-pill"><div class="num">{len(tickets)}</div><div class="lbl">待列印（人）</div></div>
          <div class="stat-pill"><div class="num">{total_count}</div><div class="lbl">總張數</div></div>
          <div class="stat-pill"><div class="num">{n_sessions}</div><div class="lbl">場次數</div></div>
          <div class="stat-pill accent"><div class="num">{n_checked}</div><div class="lbl">已勾選</div></div>
          <div class="stat-pill"><div class="num">{len(skipped)}</div><div class="lbl">已略過（歷史）</div></div>
        </div>
        """, unsafe_allow_html=True)

        # 全域歸檔按鈕
        if n_checked > 0:
            col_a1, col_a2 = st.columns([3, 1])
            with col_a1:
                st.markdown(f'<div class="note-info">已勾選 <strong>{n_checked}</strong> 筆（可能跨多個場次），歸檔後下次不再出現。</div>', unsafe_allow_html=True)
            with col_a2:
                if st.button(f"✅ 歸檔勾選的 {n_checked} 筆", type="primary", use_container_width=True):
                    to_archive = [t for t in tickets if t["key"] in st.session_state.checked_keys]
                    keys   = [t["key"] for t in to_archive]
                    labels = [make_current_label(t) for t in to_archive]
                    if save_history(HISTORY_SHEET_ID, keys, labels):
                        st.session_state.history_set.update(keys)
                        st.session_state.tickets = [t for t in tickets if t["key"] not in st.session_state.checked_keys]
                        st.session_state.checked_keys = set()
                        st.success(f"✅ 已歸檔 {len(keys)} 筆！")
                        st.rerun()

        if not tickets:
            st.markdown('<div class="note-ok">🎉 沒有待列印資料，全部都已處理完成。</div>', unsafe_allow_html=True)
        else:
            grouped = group_by_session(tickets)
            session_names = list(grouped.keys())
            tabs = st.tabs([f"🗓 {s}" for s in session_names])

            for tab, session_display in zip(tabs, session_names):
                with tab:
                    sess_tickets = grouped[session_display]
                    sess_total   = sum(t["count"] for t in sess_tickets)
                    sess_checked = sum(1 for t in sess_tickets if t["key"] in st.session_state.checked_keys)
                    st.caption(f"{len(sess_tickets)} 人 ／ {sess_total} 張　已勾 {sess_checked}/{len(sess_tickets)}")

                    sheet_groups = {}
                    for t in sess_tickets:
                        sheet_groups.setdefault(t["sheet"], []).append(t)

                    for sheet_name, group_tickets in sheet_groups.items():
                        import hashlib
                        safe_id = hashlib.md5(f"{session_display}_{sheet_name}".encode("utf-8")).hexdigest()[:12]
                        editor_key = f"editor_{safe_id}"

                        sg_total   = sum(st.session_state.edits.get(t["key"], {}).get("count", t["count"]) for t in group_tickets)
                        sg_checked = sum(1 for t in group_tickets if t["key"] in st.session_state.checked_keys)

                        st.markdown(
                            f'<div class="group-head"><span>📂 {sheet_name}</span>'
                            f'<span>{len(group_tickets)} 人 ／ {sg_total} 張 ／ 已勾 {sg_checked}</span></div>',
                            unsafe_allow_html=True
                        )

                        # 組裝表格資料
                        rows = []
                        for t in group_tickets:
                            ed = st.session_state.edits.get(t["key"], {})
                            seats_default = ed.get("seats")
                            if seats_default is None:
                                seats_default = "、".join(t.get("seats", []))
                            rows.append({
                                "已列印": t["key"] in st.session_state.checked_keys,
                                "編號":   t["id"],
                                "姓名":   ed.get("name", t["name"]),
                                "電話":   ed.get("tel", t.get("tel", "")),
                                "座位":   seats_default,
                                "張數":   int(ed.get("count", t["count"])),
                            })
                        base_df = pd.DataFrame(rows)

                        edited_df = st.data_editor(
                            base_df,
                            column_config={
                                "已列印": st.column_config.CheckboxColumn("已列印", width="small"),
                                "編號":   st.column_config.TextColumn("編號", width="small", disabled=True),
                                "姓名":   st.column_config.TextColumn("姓名", width="medium"),
                                "電話":   st.column_config.TextColumn("電話", width="medium"),
                                "座位":   st.column_config.TextColumn("座位", width="large"),
                                "張數":   st.column_config.NumberColumn("張數", width="small", min_value=1, max_value=300),
                            },
                            hide_index=True,
                            use_container_width=True,
                            height=min(440, 38 * (len(group_tickets) + 1) + 3),
                            key=editor_key,
                        )

                        # 同步編輯結果回 session_state
                        for i, t in enumerate(group_tickets):
                            row = edited_df.iloc[i]
                            if bool(row["已列印"]):
                                st.session_state.checked_keys.add(t["key"])
                            else:
                                st.session_state.checked_keys.discard(t["key"])

                            new_edit = {}
                            if str(row["姓名"]) != str(t["name"]):
                                new_edit["name"] = row["姓名"]
                            if int(row["張數"]) != int(t["count"]):
                                new_edit["count"] = int(row["張數"])
                            orig_tel = t.get("tel", "")
                            if str(row["電話"]) != str(orig_tel):
                                new_edit["tel"] = row["電話"]
                            orig_seats = "、".join(t.get("seats", []))
                            if str(row["座位"]) != orig_seats:
                                new_edit["seats"] = row["座位"]
                            if new_edit:
                                st.session_state.edits.setdefault(t["key"], {}).update(new_edit)

                        # 操作列：全選／取消全選／複製
                        gc1, gc2, gc3 = st.columns([1, 1, 2])
                        with gc1:
                            if st.button("✅ 全選", key=f"selall_{editor_key}", use_container_width=True):
                                for t in group_tickets:
                                    st.session_state.checked_keys.add(t["key"])
                                st.session_state.pop(editor_key, None)
                                st.rerun()
                        with gc2:
                            if st.button("☐ 取消全選", key=f"unselall_{editor_key}", use_container_width=True):
                                for t in group_tickets:
                                    st.session_state.checked_keys.discard(t["key"])
                                st.session_state.pop(editor_key, None)
                                st.rerun()
                        with gc3:
                            unprinted = [t for t in group_tickets if t["key"] not in st.session_state.checked_keys]
                            if unprinted:
                                with st.popover(f"📋 複製未列印（{len(unprinted)} 筆）", use_container_width=True):
                                    st.code("\n".join(make_current_label(t) for t in unprinted), language=None)

        # 下載修改後的 xlsx
        if st.session_state.edits and st.session_state.get("uploaded_file_bytes"):
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">📥 下載修改後的 xlsx</div>', unsafe_allow_html=True)
            st.caption(f"已修改 {len(st.session_state.edits)} 筆資料，下載後可取代原始檔案")
            if st.button("產生修改後的 xlsx"):
                try:
                    updated_bytes = apply_edits_to_xlsx(
                        st.session_state.uploaded_file_bytes,
                        st.session_state.selected_sheets[0] if st.session_state.selected_sheets else None,
                        st.session_state.edits,
                        st.session_state.tickets + st.session_state.skipped
                    )
                    st.download_button(
                        "⬇️ 點此下載",
                        data=updated_bytes,
                        file_name=f"修改後報名表_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"產生失敗：{e}")
            st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# 頁面：📋 簽到表
# ══════════════════════════════════════════════════════════
elif st.session_state.page == "signin":

    if not st.session_state.rule_confirmed or not st.session_state.selected_sheets:
        st.markdown('<div class="note-info">請先到「📤 上傳與處理」完成上傳與規則確認。</div>', unsafe_allow_html=True)
    else:
        sname = st.session_state.selected_sheets[0]
        row0_str = " ".join(str(c) for c in st.session_state.raw_sheets[sname].iloc[0].tolist())
        is_member = ("姓名" in row0_str and "張數" in row0_str and "座位" in row0_str)

        if not is_member:
            st.markdown('<div class="note-warn">此工作表格式尚未支援產生簽到表，請聯絡管理員更新程式。</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">簽到表設定 ＆ 下載</div>', unsafe_allow_html=True)
            col_name, col_btn = st.columns([4, 1])
            with col_name:
                show_name_input = st.text_input(
                    "演出名稱（顯示在簽到表標題）",
                    value="親子音樂劇《阿甯咕的爸鼻不見了？》",
                    key="show_name_input"
                )
            with col_btn:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.session_state.get("uploaded_file_bytes"):
                    excel_bytes = generate_signin_excel(st.session_state.uploaded_file_bytes, sname, show_name_input)
                    st.download_button(
                        "📥 下載 Excel", data=excel_bytes,
                        file_name=f"簽到表_{sname}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary"
                    )
                else:
                    st.warning("請重新上傳 xlsx")
            st.markdown('</div>', unsafe_allow_html=True)

            all_tickets = st.session_state.tickets + st.session_state.skipped
            if all_tickets:
                preview_sessions = defaultdict(list)
                for t in sorted(all_tickets, key=lambda x: (x["earliest_sort"], int(x["id"]))):
                    preview_sessions[t["earliest_display"]].append(t)

                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<div class="card-title">👁 簽到表預覽</div>', unsafe_allow_html=True)
                session_names = sorted(preview_sessions.keys(), key=lambda s: preview_sessions[s][0]["earliest_sort"])
                tabs = st.tabs([f"🗓 {s}" for s in session_names])
                for tab, session_display in zip(tabs, session_names):
                    with tab:
                        sess_tickets = preview_sessions[session_display]
                        sess_total = sum(t["count"] for t in sess_tickets)
                        st.caption(f"{len(sess_tickets)} 人 ／ {sess_total} 張")
                        preview_data = [{
                            "編號": t["id"], "社群帳號": t.get("sns",""), "姓名": t["name"],
                            "電話": t.get("tel",""), "座位": "　".join(t.get("seats",[])),
                            "張數": t["count"], "領取簽名": ""
                        } for t in sess_tickets]
                        st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True,
                                     height=min(500, max(150, len(sess_tickets) * 36 + 40)))
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="note-info">請先完成標籤產生，才能預覽簽到表。</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# 頁面：🖨️ 補印
# ══════════════════════════════════════════════════════════
elif st.session_state.page == "reprint":

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">🖨️ 補印（不受歷史紀錄限制，可重複產生）</div>', unsafe_allow_html=True)
    st.caption("輸入編號重新產生標籤，例如：4 或 4,7,12")

    col_in, col_btn = st.columns([4, 1])
    with col_in:
        reprint_id = st.text_input("編號", placeholder="例如：4 或 4,7,12", key="reprint_input", label_visibility="collapsed")
    with col_btn:
        do_reprint = st.button("產生補印標籤", type="primary", use_container_width=True)

    if do_reprint and reprint_id:
        tickets_now = st.session_state.get("tickets", []) + st.session_state.get("skipped", [])
        ids = [x.strip() for x in reprint_id.replace("，", ",").split(",") if x.strip()]
        found = [t for t in tickets_now if t["id"] in ids]
        if found:
            st.code("\n".join(make_current_label(t) for t in found), language=None)
        else:
            st.markdown(f'<div class="note-warn">找不到編號：{", ".join(ids)}（請先完成上傳與標籤產生）</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# 頁面：⚙️ 歷史與設定
# ══════════════════════════════════════════════════════════
elif st.session_state.page == "settings":

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📦 歷史紀錄</div>', unsafe_allow_html=True)
    col_h1, col_h2, col_h3 = st.columns([2, 1, 1])
    with col_h1:
        st.markdown(f'<div class="stat-pill"><div class="num">{len(st.session_state.history_set)}</div><div class="lbl">已列印歷史筆數</div></div>', unsafe_allow_html=True)
    with col_h2:
        if st.button("🔄 重新載入歷史", use_container_width=True):
            with st.spinner("連線中..."):
                st.session_state.history_set = load_history(HISTORY_SHEET_ID)
            st.success(f"已載入 {len(st.session_state.history_set)} 筆")
    with col_h3:
        if st.button("🗑️ 重新開始（清除本次資料）", use_container_width=True):
            for k in ["raw_sheets","selected_sheets","rule_confirmed","tickets","skipped",
                      "warnings","checked_keys","edits","uploaded_file_bytes"]:
                st.session_state[k] = defaults[k]
            st.success("已清除，請重新上傳")
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    skipped = st.session_state.skipped
    if skipped:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f'<div class="card-title">⏭️ 已列印過的資料（{len(skipped)} 筆，可取消歸檔重新列印）</div>', unsafe_allow_html=True)
        for t in skipped:
            sk1, sk2 = st.columns([1, 14])
            with sk1:
                if st.button("↩️", key=f"unarchive_{t['key']}", use_container_width=True, help="取消歸檔，恢復成待列印"):
                    if remove_history(HISTORY_SHEET_ID, [t["key"]]):
                        st.session_state.history_set.discard(t["key"])
                        entry = dict(t)
                        entry["is_new"] = True
                        st.session_state.tickets.append(entry)
                        st.session_state.tickets.sort(key=lambda x: (x["earliest_sort"], int(x["id"])))
                        st.session_state.skipped = [s for s in st.session_state.skipped if s["key"] != t["key"]]
                        st.success(f"已取消歸檔：{t['label']}")
                        st.rerun()
            with sk2:
                st.markdown(f'<span style="color:#aaa;font-size:0.83rem;"><s>{t["label"]}</s></span>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="note-info">目前沒有本次已列印過的資料紀錄可顯示。</div>', unsafe_allow_html=True)
